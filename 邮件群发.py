"""
Email sender — reads Excel recipient list, sends Word doc as email body via MailEnvelope.
Double-click .exe to run after placing it alongside the .xlsm file.

Requires: Windows + Word + Outlook (no Python needed when packaged as exe)
Build:   pip install pyinstaller pywin32 openpyxl
         pyinstaller --onefile --windowed --name "邮件群发" mail_sender.py
"""
import sys
import os
import time
from pathlib import Path

# ── Frozen exe: chdir to exe location so relative paths work ──
if getattr(sys, 'frozen', False):
    # sys.argv[0] = the .exe the user double-clicked
    # sys.executable = temp extraction path (NOT the exe dir!)
    exe_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
    os.chdir(exe_dir)

# ═══════════════════════════════════════════════════════════════
# Configuration — edit these to match your Excel structure
# ═══════════════════════════════════════════════════════════════
EXCEL_FILE = "邮件模板.xlsm"       # or .xlsx — the workbook with sheets below
SUPPLIER_SHEET = "Supplier List"   # columns: A=To, B=CC1, C=CC2, D=CC3
CONTENT_SHEET = "Content"          # B1=subject, B2=body text (hyperlink → Word doc)
ROWS_START = 2                     # data starts on this row (1 = header)

def show_dialog(title, msg):
    """Show a Windows message box (works in exe mode)."""
    try:
        import ctypes
        ctypes.windll.user32.MessageBoxW(0, str(msg), str(title), 0x40)
    except Exception:
        print(f"[{title}] {msg}")


def read_excel(path):
    """Return (recipients, subject, body_text, word_doc_path)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)

    # ── Supplier List ──
    if SUPPLIER_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{SUPPLIER_SHEET}' not found in {path}")
    ws_sup = wb[SUPPLIER_SHEET]

    recipients = []
    for row in ws_sup.iter_rows(min_row=ROWS_START,
                                max_row=ws_sup.max_row,
                                max_col=4, values_only=True):
        to_addr = str(row[0]).strip() if row[0] else ""
        if not to_addr or to_addr == "0" or to_addr == "None":
            continue
        cc_parts = []
        for i in range(1, 4):
            val = str(row[i]).strip() if len(row) > i and row[i] else ""
            if val and val != "0" and val != "None":
                cc_parts.append(val)
        recipients.append((to_addr, ";".join(cc_parts)))

    # ── Content ──
    if CONTENT_SHEET not in wb.sheetnames:
        raise ValueError(f"Sheet '{CONTENT_SHEET}' not found in {path}")
    ws_cnt = wb[CONTENT_SHEET]

    subject = str(ws_cnt.cell(1, 2).value or "").strip()
    body_text = str(ws_cnt.cell(2, 2).value or "").strip()

    # Resolve Word doc hyperlink from cell B2
    word_doc_path = ""
    cell = ws_cnt.cell(2, 2)
    if cell.hyperlink and cell.hyperlink.target:
        word_doc_path = cell.hyperlink.target
        # Resolve relative path against Excel file's directory
        if ":" not in word_doc_path:
            base_dir = os.path.dirname(os.path.abspath(path))
            word_doc_path = os.path.join(base_dir, word_doc_path)
    wb.close()

    if not word_doc_path or not Path(word_doc_path).exists():
        raise FileNotFoundError(f"Word document not found: {word_doc_path}")

    return recipients, subject, body_text, word_doc_path


def send_via_word(doc_path, subject, to_addr, cc_str, attachments):
    """Send email via Word MailEnvelope. Returns True on success."""
    import pythoncom
    pythoncom.CoInitialize()  # needed in threads / exe

    try:
        import win32com.client as win32
    except ImportError:
        # Fallback for environments without early-bound win32com
        import comtypes.client
        word = comtypes.client.CreateObject("Word.Application")
    else:
        word = win32.Dispatch("Word.Application")

    try:
        word.Visible = False
        doc = word.Documents.Open(doc_path, ReadOnly=True)

        # MailEnvelope: the document IS the email body
        doc.MailEnvelope.Introduction = ""
        mail = doc.MailEnvelope.Item

        mail.Subject = subject
        mail.To = to_addr
        if cc_str:
            mail.CC = cc_str

        if attachments:
            for att in attachments:
                mail.Attachments.Add(att)

        mail.Send()
        time.sleep(2)  # wait for async send to complete
        doc.Close(False)
        word.Quit()
        pythoncom.CoUninitialize()
        return True

    except Exception:
        try:
            doc.Close(False)
        except Exception:
            pass
        try:
            word.Quit()
        except Exception:
            pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
        raise


def send_plain_text(subject, body_text, to_addr, cc_str, attachments):
    """Fallback: plain text via Outlook."""
    import pythoncom
    pythoncom.CoInitialize()

    try:
        import win32com.client as win32
    except ImportError:
        import comtypes.client
        outlook = comtypes.client.CreateObject("Outlook.Application")
    else:
        outlook = win32.Dispatch("Outlook.Application")

    mail = outlook.CreateItem(0)
    mail.Subject = subject
    mail.Recipients.Add(f"SMTP:{to_addr}").Type = 1  # olTo
    if cc_str:
        for addr in cc_str.split(";"):
            addr = addr.strip()
            if addr:
                mail.Recipients.Add(f"SMTP:{addr}").Type = 2  # olCC
    mail.Body = body_text

    if attachments:
        for att in attachments:
            mail.Attachments.Add(att)

    mail.Send()
    return True


def main():
    excel_path = EXCEL_FILE
    if not Path(excel_path).exists():
        show_dialog("错误", f"找不到 Excel 文件:\n{excel_path}\n\n请把 exe 和 {EXCEL_FILE} 放在同一文件夹。")
        sys.exit(1)

    try:
        recipients, subject, body_text, doc_path = read_excel(excel_path)
    except Exception as e:
        show_dialog("读取失败", f"读取 Excel 出错:\n{e}")
        sys.exit(1)

    # ── Optional: pick file attachments ──
    attachments = []
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        files = filedialog.askopenfilenames(
            title="选择附加文件（可跳过）",
            filetypes=[("所有文件", "*.*")]
        )
        root.destroy()
        if files:
            attachments = list(files)
    except Exception:
        pass  # no tkinter in some Python builds — skip

    # ── Send loop ──
    total = len(recipients)
    sent = 0
    failed = 0
    log_lines = [f"=== {time.strftime('%Y-%m-%d %H:%M:%S')} ===",
                 f"Subject: {subject}",
                 f"Word doc: {doc_path}",
                 f"Recipients: {total}\n"]

    for idx, (to_addr, cc_str) in enumerate(recipients, 1):
        line = f"[{idx}/{total}] {to_addr}"
        if cc_str:
            line += f"  CC: {cc_str}"
        print(line)

        try:
            ok = send_via_word(doc_path, subject, to_addr, cc_str, attachments)
            if ok:
                sent += 1
                log_lines.append(f"OK  [{idx}] {to_addr}")
            else:
                # MailEnvelope failed — fall back to plain text
                send_plain_text(subject, body_text, to_addr, cc_str, attachments)
                sent += 1
                log_lines.append(f"OK* [{idx}] {to_addr} (plain text)")
        except Exception as e:
            failed += 1
            log_lines.append(f"FAIL [{idx}] {to_addr} — {e}")
            print(f"  FAIL: {e}")

    # ── Summary ──
    elapsed = time.strftime("%M:%S", time.gmtime(time.perf_counter()))
    summary = f"\nSent: {sent}  Failed: {failed}  Time: {elapsed}\n"
    log_lines.append(summary)
    print(summary)

    # Write log
    log_path = Path(excel_path).parent / "发送日志.txt"
    with open(log_path, "w", encoding="utf-8") as f:
        f.write("\n".join(log_lines))

    show_dialog("完成", f"发送完毕！\n\n成功: {sent}\n失败: {failed}\n日志: 发送日志.txt")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        traceback.print_exc()
        show_dialog("出错", f"程序异常:\n{e}")
